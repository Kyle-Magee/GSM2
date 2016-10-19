from openpyxl import load_workbook
from workers import Employee, Parttime, ParttimeCashier, ParttimePricer, Pricer, Cashier, Manager


def excel_sheet(filename):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')
    employee_list = {}

    def read_sheet():
        """Takes in the values for all names in the left-hand column and the weekly-budget"""
        for row_num in range(1, 50):
            entry = str(sheet.cell(row=row_num, column=1).value)
            if entry.isupper():
                availability = {}
                job_position = sheet.cell(row=row_num, column=6).value
                for (day_columns, day) in zip(range(4, 25, 3), ('sunday', 'monday', 'tuesday', 'wednesday', 'thursday',
                                                                'friday')):
                    am_i_free = str(sheet.cell(row=row_num, column=day_columns).value)
                    if am_i_free == 'X':
                        availability[day] = False
                    else:
                        availability[day] = True

                create_object(row_num, job_position, availability)

        employee_list['week_budget'] = int(sheet.cell(row=2, column=13).value)

    def create_object(row_num, job_position, availability):
        id_row = str(row_num)
        if row_num < 13:
            employee_list[row_num] = Manager(id_row, **availability)
        elif 13 < row_num < 26:
            if job_position == 'F':
                employee_list[row_num] = Cashier(row_num, **availability)
            elif job_position == 'B':
                employee_list[row_num] = Pricer(row_num, **availability)
            else:
                employee_list[row_num] = Employee(row_num, **availability)
        elif row_num > 26:
            if job_position == 'F':
                employee_list[row_num] = ParttimeCashier(row_num, **availability)
            elif job_position == 'B':
                employee_list[row_num] = ParttimePricer(row_num, **availability)
            else:
                employee_list[row_num] = Parttime(row_num, **availability)

    read_sheet()
    return employee_list
