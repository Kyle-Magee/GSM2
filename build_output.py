from take_input import excel_sheet
from build_schedule import make_a_schedule
from translate_for_output import translate_for_output
from openpyxl import load_workbook


def build_output(folder, filename, startdate, template_file):
    schedule = make_a_schedule(template_file)
    schedule = translate_for_output(schedule)
    filename = folder + '/' + filename
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')

    def write_shifts():
        for day in schedule:
            for shift in schedule[day].shifts:
                for (index, tuple) in enumerate(schedule[day].shifts[shift]):
                    try:
                        row_num = schedule[day].shifts[shift][index][0]
                        hours = schedule[day].shifts[shift][index][-1]
                        if hours == 8:
                            hours = 7.5
                        start_time = shift[0]
                        end_time = shift[-1]
                        sheet.cell(row=row_num, column=day).value = start_time
                        sheet.cell(row=row_num, column=day+1).value = end_time
                        sheet.cell(row=row_num, column=day+2).value = hours
                    except IndexError:
                        continue

    def write_off_days():
        for row_num in range(8, 50):
            for col_num in range(4, 29, 3):
                if sheet.cell(row=row_num, column=col_num).value in ('X', '#'):
                    sheet.cell(row=row_num, column=col_num).value = 'OFF'
                    sheet.cell(row=row_num, column=col_num + 1).value = ''
                    sheet.cell(row=row_num, column=col_num + 2).value = 0

    def write_date():
        sheet.cell(row=4, column=4).value = startdate

    def save_to_file():
        formatted_startdate = startdate.replace('/', '-')
        new_filename = "002.xlsx"
        wb.save(new_filename)

    write_shifts()
    write_off_days()
    write_date()
    save_to_file()

